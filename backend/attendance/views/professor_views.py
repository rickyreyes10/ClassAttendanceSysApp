from django.shortcuts import render
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.db.utils import IntegrityError
from django.http import JsonResponse, HttpResponse
import os
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.views.decorators.csrf import csrf_exempt
from attendance.models import Professor, Class, Student, AttendanceLog
from attendance.serializers import ProfessorSerializer, ClassSerializer
import bcrypt


@api_view(['POST'])
def professor_create_class(request):
    # Extracting data from the POST request
    data = request.data
    course_name = data.get('courseName', '').strip()
    crn = data.get('crn', '').strip()
    email = data.get('email', '').strip()
    password = data.get('password', '').strip()

    # Check if the provided CRN already exists
    if Class.objects.filter(crn=crn).exists():
        return Response({"status": "error", "error": "The provided CRN already exists! Please provide a unique CRN."}, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if professor exists with the provided email
    try:
        professor = Professor.objects.get(email=email)
        # If the professor exists, validate their password
        if not bcrypt.checkpw(password.encode('utf-8'), professor.hashed_password.encode('utf-8')):
            return Response({"status": "error", "error": "Invalid password provided!"}, status=status.HTTP_400_BAD_REQUEST)
    except Professor.DoesNotExist:
        # If the professor doesn't exist, insert the professor
        hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        professor = Professor(email=email, hashed_password=hashed_password)
        professor.save()

    # Insert the class details
    class_instance = Class(crn=crn, course_name=course_name, professor=professor)
    class_instance.save()

    return Response({"status": "success"}, status=status.HTTP_201_CREATED)


@api_view(['POST'])
def professor_login_class(request):
    data = request.data
    crn = data.get("crn", "").strip()
    email = data.get("email", "").strip()
    password = data.get("password", "").strip()

    # Check if provided professor credentials match the database records.
    try:
        # Check if class with CRN exists and matches the provided email
        matching_class = Class.objects.filter(crn=crn, professor__email=email).first()
        if matching_class:
            stored_hashed_password = matching_class.professor.hashed_password
            if professor_verify_password(password, stored_hashed_password):
                # Store the crn in the session
                request.session['crn'] = crn
                return Response({"status": "success", "message": "Login successful"}, status=status.HTTP_200_OK)
            else:
                return Response({"status": "error", "message": "Invalid password"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"status": "error", "message": "CRN or Email are invalid"}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({"status": "error", "message": str(e)}, status=status.HTTP_400_BAD_REQUEST)

def professor_verify_password(provided_password, stored_hashed_password):
    return bcrypt.checkpw(provided_password.encode('utf-8'), stored_hashed_password.encode('utf-8'))



@api_view(['GET'])
def professor_generate_attendance(request):
    # Retrieve the crn from the session
    crn = request.session.get('crn')
    if not crn:
        return Response({"status": "error", "message": "CRN not found in session. Please login again."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Fetching the attendance log for the given CRN
        attendance_data = AttendanceLog.objects.filter(class_crn=crn).values_list(
            'student__student_name', 'attendance_date', 'attendance_time', 'method'
        ).order_by('attendance_date', 'attendance_time')

        if not attendance_data:
            return Response({"status": "error", "message": "No attendance log found for this CRN."}, status=status.HTTP_404_NOT_FOUND)

        # Generating Excel Report
        workbook = Workbook()
        sheet = workbook.active

        headers = ['Student Name', 'Attendance Date', 'Attendance Time', 'Method']
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        for row_num, row_data in enumerate(attendance_data, 2):
            for col_num, cell_data in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=cell_data)

        excel_path = os.path.join("/tmp", f"{crn}_attendance_report.xlsx")  # temporarily saving to /tmp
        workbook.save(excel_path)

        with open(excel_path, 'rb') as excel_file:
            response = HttpResponse(excel_file.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{crn}_attendance_report.xlsx"'
            return response

    except Exception as e:
        return Response({"status": "error", "message": str(e)}, status=status.HTTP_400_BAD_REQUEST)   