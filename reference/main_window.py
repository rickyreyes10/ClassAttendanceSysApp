import os
import tkinter as tk
from tkinter import messagebox
from choice_window import run_choice_window
import util
import hashlib
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from db_utils import connect


class MainWindow:
    # Constructor for the MainWindow class
    def __init__(self, root):
        self.root = root
        self.frame = None
        self.show()

    # Displays the main UI elements on the screen
    def show(self):
        # Set initial geometry and title
        self.root.geometry("1200x750")
        self.root.title("Class Attendance System")

        # Destroy the previous frame if it exists
        if self.frame:
            self.frame.destroy()

        # Create a new frame
        self.frame = tk.Frame(self.root, bg='black')
        self.frame.pack(expand=True, fill="both")

        # Display label and buttons on the main window
        label = util.get_text_label(self.frame, "Class Attendance System", font_size=32, justify="center",
                                    fg_color='white', bg_color='black')
        label.pack(pady=20)

        # Professor button
        professor_button = util.get_button(self.frame, "Professor", color="#0066cc", command=self.professor_options,
                                           font_size=30, height=4, width=40)
        professor_button.pack(pady=10)

        # Student button
        student_button = util.get_button(self.frame, "Student", color="#009966", command=self.student_choice,
                                         font_size=30, height=4, width=40)
        student_button.pack(pady=10)

    # Displays the professor options UI
    def professor_options(self):
        # Display professor options in a new Toplevel window
        self.option_window = tk.Toplevel(self.root)
        self.option_window.title("Professor Options")

        # Label for guidance
        label = util.get_text_label(self.option_window, text="Select an option:")
        label.pack(pady=10)

        # Login and Create buttons for professors
        # Login button
        login_btn = util.get_button(
            window=self.option_window,
            text="Login",
            color="blue",
            command=self.on_login_click,
            font_size=20
        )
        login_btn.pack(pady=10)

        # Create button
        create_btn = util.get_button(
            window=self.option_window,
            text="Create",
            color="green",
            command=self.on_create_click,
            font_size=20
        )
        create_btn.pack(pady=10)

    # Redirects the UI to the professor login screen
    # Handle click on the Login button
    def on_login_click(self):
        self.option_window.destroy()  # Close the options window
        self.professor_login()

    # Redirects the UI to the class creation screen
    # Handle click on the Create button
    def on_create_click(self):
        self.option_window.destroy()  # Close the options window
        self.create_class()

    # Displays the login form for professors
    # Display the professor login page
    def professor_login(self):
        self.root.geometry("1200x750")
        self.root.title("Professor Login")

        if self.frame:
            self.frame.destroy()

        self.frame = tk.Frame(self.root, bg='black')
        self.frame.pack(expand=True, fill="both")

        # Input fields for CRN, email, and password
        # CRN Entry
        crn_label = util.get_text_label(self.frame, "Enter CRN", font_size=24, justify="left", fg_color='white',
                                        bg_color='black')
        crn_label.pack(pady=10)
        self.crn_entry = util.get_entry_text(self.frame)
        self.crn_entry.pack(pady=10)

        # Email Entry
        email_label = util.get_text_label(self.frame, "Enter Email", font_size=24, justify="left", fg_color='white',
                                          bg_color='black')
        email_label.pack(pady=10)
        self.email_entry = util.get_entry_text(self.frame)
        self.email_entry.pack(pady=10)

        # Password Entry
        password_label = util.get_text_label(self.frame, "Enter Password", font_size=24, justify="left",
                                             fg_color='white', bg_color='black')
        password_label.pack(pady=10)
        self.password_entry = util.get_entry_text(self.frame, height=1)
        self.password_entry.pack(pady=10)

        # Login button
        login_button = util.get_button(self.frame, "Login", color="#009966", command=self.validate_login, font_size=20)
        login_button.pack(pady=30)

    # Displays the form to create a new class
    # Display the class creation page for professors
    def create_class(self):
        self.root.geometry("1200x750")
        self.root.title("Create New Class")

        if self.frame:
            self.frame.destroy()

        self.frame = tk.Frame(self.root, bg='black')
        self.frame.pack(expand=True, fill="both")

        # Input fields for course name, CRN, email, and password
        # Course Name Entry
        course_label = util.get_text_label(self.frame, "Enter Course Name", font_size=24, justify="left",
                                           fg_color='white', bg_color='black')
        course_label.pack(pady=10)
        self.course_name_entry = util.get_entry_text(self.frame)
        self.course_name_entry.pack(pady=10)

        # CRN Entry
        crn_label = util.get_text_label(self.frame, "Enter CRN", font_size=24, justify="left", fg_color='white',
                                        bg_color='black')
        crn_label.pack(pady=10)
        self.crn_entry = util.get_entry_text(self.frame)
        self.crn_entry.pack(pady=10)

        # Email Entry
        email_label = util.get_text_label(self.frame, "Enter Email", font_size=24, justify="left", fg_color='white',
                                          bg_color='black')
        email_label.pack(pady=10)
        self.email_entry = util.get_entry_text(self.frame)
        self.email_entry.pack(pady=10)

        # Password Entry
        password_label = util.get_text_label(self.frame, "Enter Password", font_size=24, justify="left",
                                             fg_color='white', bg_color='black')
        password_label.pack(pady=10)
        self.password_entry = util.get_entry_text(self.frame, height=1)
        self.password_entry.pack(pady=10)

        # Create button
        create_button = util.get_button(self.frame, "Create", color="#009966", command=self.save_class_details,
                                        font_size=20)
        create_button.pack(pady=30)

    # Displays the CRN input form for students
    # Display the CRN input page for students
    def student_choice(self):
        self.root.geometry("800x500")
        self.root.title("Student Course Entry")

        if self.frame:
            self.frame.destroy()

        self.frame = tk.Frame(self.root, bg='black')
        self.frame.pack(expand=True, fill='both')

        # CRN entry for the student
        crn_label = util.get_text_label(self.frame, "Enter Your CRN", font_size=24, justify="left",
                                        fg_color='white', bg_color='black')
        crn_label.pack(pady=10)
        self.student_crn_entry = util.get_entry_text(self.frame)
        self.student_crn_entry.pack(pady=10)

        proceed_button = util.get_button(self.frame, "Proceed", color="#009966",
                                         command=self.proceed_to_student_choice, font_size=20)
        proceed_button.pack(pady=30)

    # Validates the entered CRN and proceeds to the next step
    def proceed_to_student_choice(self):
        crn = self.student_crn_entry.get("1.0", tk.END).strip()  # Fetch the entered CRN
        if self.validate_crn(crn):  # Check if the CRN is valid
            # Pass the CRN to the next step (either face recognition or QR code or however you handle it)
            run_choice_window(self.root, self.frame, crn)
        else:
            messagebox.showerror("Error", "Please enter a valid CRN")

    # Checks if the given CRN exists in the classes table
    def validate_crn(self, crn):
        conn = connect()
        cursor = conn.cursor()

        query = "SELECT COUNT(*) FROM Classes WHERE crn = %s"
        cursor.execute(query, (crn,))
        result = cursor.fetchone()
        conn.close()

        return bool(result)  # will return True if the CRN is found, else False

    # Reads attendance data and generates an attendance log in both Excel and PDF format
    def generate_attendance_log(self, crn):
        conn = None
        attendance_data = []

        try:
            conn = connect()
            cursor = conn.cursor()

            # Fetching the attendance log for the given CRN
            query = """ 
            SELECT Students.student_name, Attendance_Log.attendance_date, Attendance_Log.attendance_time, Attendance_Log.method 
            FROM Attendance_Log 
            JOIN Students ON Attendance_Log.student_email = Students.email
            WHERE Attendance_Log.class_crn = %s
            ORDER BY Attendance_Log.attendance_date, Attendance_Log.attendance_time
            """
            cursor.execute(query, (crn,))
            attendance_data = cursor.fetchall()

            if not attendance_data:
                messagebox.showerror("Error", "No attendance log found for this CRN.")
                return

            # Generating Excel Report
            workbook = Workbook()
            sheet = workbook.active

            headers = ['Student Name', 'Attendance Date', 'Attendance Time', 'Method']
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            for row_num, row_data in enumerate(attendance_data, 2):
                for col_num, cell_data in enumerate(row_data, 1):
                    sheet.cell(row=row_num, column=col_num, value=cell_data)

            excel_path = f"./db/{crn}_attendance_report.xlsx"
            workbook.save(excel_path)

            # Generating PDF Report
            pdf_path = f"./db/{crn}_attendance_report.pdf"
            c = canvas.Canvas(pdf_path, pagesize=letter)
            width, height = letter

            # Set up the headers in the PDF
            for col_num, header in enumerate(headers, 1):
                c.drawString(100 + (150 * (col_num - 1)), height - 50, header)

            for row_num, row_data in enumerate(attendance_data, 2):
                y_position = height - 50 - (15 * row_num)  # Adjust as per your requirements
                for col_num, cell_data in enumerate(row_data, 1):
                    c.drawString(100 + (150 * (col_num - 1)), y_position, str(cell_data))

            c.save()

            messagebox.showinfo("Success", "Reports generated successfully!")

        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            if conn and conn.is_connected():
                conn.close()

    # Collects class details from the form, hashes the password and stores the details
    def save_class_details(self):
        crn = self.crn_entry.get("1.0", tk.END).strip()
        course_name = self.course_name_entry.get("1.0", tk.END).strip()
        email = self.email_entry.get("1.0", tk.END).strip()
        password = self.password_entry.get("1.0", tk.END).strip()

        hashed_password = self.hash_password(password)

        conn = None
        try:
            conn = connect()
            cursor = conn.cursor()

            # Check if professor exists with the provided email
            prof_query = "SELECT professor_id FROM Professors WHERE email = %s"
            cursor.execute(prof_query, (email,))
            prof_result = cursor.fetchone()

            # if not, insert the professor
            if not prof_result:
                insert_prof_query = """
                INSERT INTO Professors (email, hashed_password) VALUES (%s, %s)
                """
                cursor.execute(insert_prof_query, (email, hashed_password))
                conn.commit()

            # Insert the class details
            insert_class_query = """
            INSERT INTO Classes (crn, course_name, professor_email) VALUES (%s, %s, %s)
            """

            cursor.execute(insert_class_query, (crn, course_name, email))
            conn.commit()

            messagebox.showinfo("Success", "Class details saved successfully!")

        except Exception as e:
            messagebox.showerror("Error", str(e))

        finally:
            if conn and conn.is_connected():
                conn.close()

    # Displays the professor dashboard UI with various options like generating attendance logs
    def show_professor_dashboard(self, crn):
        self.root.geometry("1200x750")
        self.root.title("Professor Dashboard")

        if self.frame:
            self.frame.destroy()

        self.frame = tk.Frame(self.root, bg='black')
        self.frame.pack(expand=True, fill="both")

        welcome_label = util.get_text_label(self.frame, f"Welcome, Professor for CRN {crn}", font_size=24,
                                            justify="center",
                                            fg_color='white', bg_color='black')
        welcome_label.pack(pady=20)

        generate_button = util.get_button(self.frame, "Generate Attendance Log", color="#009966",
                                          command=lambda: self.generate_attendance_log(crn), font_size=20)
        generate_button.pack(pady=30)

    # Validates the credentials entered by the professor and logs them in if they're correct
    def validate_login(self):
        crn = self.crn_entry.get("1.0", tk.END).strip()
        email = self.email_entry.get("1.0", tk.END).strip()
        password = self.password_entry.get("1.0", tk.END).strip()

        print(f"Trying to validate: CRN={crn}, Email={email}, Password={password}")

        if self.check_professor_credentials(crn, email, password):
            self.show_professor_dashboard(crn)
        else:
            messagebox.showerror("Error", "Invalid Credentials")

    # Checks the given professor's credentials against the saved data
    def check_professor_credentials(self, crn, email, password):

        conn = None
        try:
            conn = connect()
            cursor = conn.cursor()

            # check if class with CRN exists and matches the provided email
            query = """
            SELECT p.hashed_password 
            FROM Classes c 
            INNER JOIN Professors p ON c.professor_email = p.email 
            WHERE c.crn = %s AND p.email = %s
            """

            cursor.execute(query, (crn, email))
            result = cursor.fetchone()

            # If a matching record is found and the hashed passwords match, return True
            if result and self.hash_password(password) == result[0]:
                return True
            return False

        except Exception as e:
            print(str(e))
            return False

        finally:
            if conn and conn.is_connected():
                conn.close()

    # Returns the hashed version of the given password
    def hash_password(self, password):
        return hashlib.sha256(password.encode()).hexdigest()
